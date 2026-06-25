from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q
from .models import VideoMetadata
from .serializers import VideoMetadataSerializer
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

# ==========================================
# VISTAS ORIGINALES (RODRIGO)
# ==========================================

class VideoListView(generics.ListCreateAPIView):
    serializer_class = VideoMetadataSerializer
    def get_queryset(self):
        queryset = VideoMetadata.objects.all()
        tipo = self.request.query_params.get('tipo')
        search = self.request.query_params.get('search')
        if tipo: queryset = queryset.filter(tipo__iexact=tipo)
        filtros = ['usuario', 'mapa', 'etnia', 'estilizado', 'aceptado', 'genero', 'especie', 'camara', 'mateo_miguel']
        for f in filtros:
            val = self.request.query_params.get(f)
            if val: queryset = queryset.filter(**{f"{f}__icontains": val})
        if search:
            queryset = queryset.filter(Q(video_id__icontains=search) | Q(usuario__icontains=search))
        return queryset.order_by("-id")

class VideoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = VideoMetadata.objects.all()
    serializer_class = VideoMetadataSerializer

class FilterOptionsView(APIView):
    def get(self, request):
        tipo = request.query_params.get('tipo', 'censo').lower()
        fields_to_filter = ["usuario", "mapa", "genero", "etnia", "camara", "especie", "estilizado", "aceptado"]
        options = {}
        
        for field_name in fields_to_filter:
            distinct_values = VideoMetadata.objects.values_list(field_name, flat=True).distinct()
            cleaned_values = [str(v).strip() for v in distinct_values if v and str(v).lower() != 'nan']
            
            # LÓGICA DE USUARIOS DINÁMICA
            if field_name == 'usuario':
                censo_users = set(
                    str(v).strip().lower() 
                    for v in VideoMetadata.objects.filter(tipo='censo').values_list('usuario', flat=True).distinct()
                    if v and str(v).lower() != 'nan'
                )
                if tipo == 'censo':
                    cleaned_values = [v for v in cleaned_values if v.lower() in censo_users]
                else:
                    cleaned_values = [v for v in cleaned_values if v.lower() not in censo_users]
            
            options[field_name] = sorted(list(set(cleaned_values)))
        return Response(options)

@method_decorator(csrf_exempt, name='dispatch')
class AutoRegisterVideoView(APIView):
    def post(self, request):
        serializer = VideoMetadataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ==========================================
# VISTAS DE RESÚMENES Y DASHBOARD (BRUNO)
# ==========================================

class CensoSummaryView(APIView):
    permission_classes = []
    
    def get(self, request):
        import csv
        import os
        from django.conf import settings
        
        csv_path = os.path.join(settings.BASE_DIR, 'censo.csv')
        
        if not os.path.exists(csv_path):
            return Response({
                "kpis": {
                    "totalVideos": 0,
                    "humanos": 0,
                    "animales": 0,
                    "sinClasificar": 0,
                    "duracionMedia": "0s",
                    "totalMapas": 0,
                    "videosMateo": 0,
                    "videosMiguel": 0,
                    "totalSegundos": 0
                },
                "especiesData": [],
                "inicialMapas": [],
                "exclusivasMateo": [],
                "exclusivasMiguel": [],
                "generoData": [],
                "etniaData": [],
                "cruceGenEtnia": [],
                "mapasSoloHombres": [],
                "camaraData": [],
                "duracionData": [],
                "actionPlan": []
            })
            
        normalized_rows = []
        with open(csv_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                c_row = {str(k).strip().lower(): str(v).strip() for k, v in row.items() if k}
                # Se verifica si hay al menos algunos campos de identificación para ignorar filas vacías
                if not any(c_row.values()):
                    continue
                normalized_rows.append(c_row)
                
        total_videos = len(normalized_rows)
        humanos = sum(1 for r in normalized_rows if r.get("especie", "").lower() == "humano")
        animales = sum(1 for r in normalized_rows if r.get("especie", "").lower() == "animal")
        sin_clasificar = total_videos - humanos - animales

        durations = []
        for r in normalized_rows:
            d_val = r.get("duracion", "")
            if d_val:
                try:
                    durations.append(float(d_val))
                except ValueError:
                    pass

        total_segundos = sum(durations)
        duracion_media = f"{total_segundos / len(durations):.1f}s" if durations else "0.0s"

        unique_maps = sorted(list(set(r.get("mapa", "") for r in normalized_rows if r.get("mapa", ""))))
        total_mapas = len(unique_maps)

        # Obtener todos los usuarios únicos dinámicamente
        unique_users = sorted(list(set(r.get("usuario", "").strip().upper() for r in normalized_rows if r.get("usuario", ""))))
        
        videos_por_usuario = {}
        for u in unique_users:
            videos_por_usuario[u] = sum(1 for r in normalized_rows if r.get("usuario", "").strip().upper() == u)

        kpis = {
            "totalVideos": total_videos,
            "humanos": humanos,
            "animales": animales,
            "sinClasificar": sin_clasificar,
            "duracionMedia": duracion_media,
            "totalMapas": total_mapas,
            "videosMateo": videos_por_usuario.get("MATEO", 0),  # Para compatibilidad legacy
            "videosMiguel": videos_por_usuario.get("MIGUEL", 0),  # Para compatibilidad legacy
            "videosPorUsuario": videos_por_usuario,  # Estructura 100% dinámica
            "totalSegundos": int(total_segundos)
        }

        # Estadísticas de especies
        total_for_pct = total_videos if total_videos > 0 else 1
        especies_data = [
            { "label": "Humano", "value": humanos, "percent": round((humanos / total_for_pct) * 100, 1), "color": "var(--neon-cyan)" },
            { "label": "Animal", "value": animales, "percent": round((animales / total_for_pct) * 100, 1), "color": "#2ecc71" },
            { "label": "Sin Clasificar", "value": sin_clasificar, "percent": round((sin_clasificar / total_for_pct) * 100, 1), "color": "#ff4b2b" }
        ]

        # Estadísticas de Mapas dinámicas
        map_stats = {}
        for m in unique_maps:
            map_stats[m] = {
                "name": m,
                "total": 0,
                "human": 0,
                "animal": 0,
                "unclassified": 0,
                "usuarios": { u: 0 for u in unique_users },
                "genders": set(),
                "ethnicities": set()
            }

        for r in normalized_rows:
            m = r.get("mapa", "")
            if not m:
                continue
            stats = map_stats[m]
            stats["total"] += 1
            
            esp = r.get("especie", "").lower()
            if esp == "humano":
                stats["human"] += 1
            elif esp == "animal":
                stats["animal"] += 1
            else:
                stats["unclassified"] += 1
                
            u = r.get("usuario", "").strip().upper()
            if u in stats["usuarios"]:
                stats["usuarios"][u] += 1
                
            gen = r.get("genero", "").lower()
            if gen in ["hombre", "mujer"]:
                stats["genders"].add(gen)
                
            etn = r.get("etnia", "").lower()
            if etn in ["blanco", "moreno"]:
                stats["ethnicities"].add(etn)

        # Ordenar mapas por total descendente
        inicial_mapas = []
        for m in unique_maps:
            s = map_stats[m]
            inicial_mapas.append({
                "name": s["name"],
                "total": s["total"],
                "human": s["human"],
                "animal": s["animal"],
                "unclassified": s["unclassified"],
                "mateo": s["usuarios"].get("MATEO", 0),
                "miguel": s["usuarios"].get("MIGUEL", 0),
                "usuarios": s["usuarios"]
            })
        inicial_mapas.sort(key=lambda x: x["total"], reverse=True)

        # Exclusivas por usuario dinámico
        exclusivas_por_usuario = { u: [] for u in unique_users }
        for m in unique_maps:
            s = map_stats[m]
            active_users = [u for u, count in s["usuarios"].items() if count > 0]
            if len(active_users) == 1:
                exclusive_user = active_users[0]
                exclusivas_por_usuario[exclusive_user].append({ "name": m, "total": s["total"] })
                
        for u in unique_users:
            exclusivas_por_usuario[u].sort(key=lambda x: x["name"])

        exclusivas_mateo = exclusivas_por_usuario.get("MATEO", [])
        exclusivas_miguel = exclusivas_por_usuario.get("MIGUEL", [])

        # Estadisticas de genero
        gender_counts = {"Hombre": 0, "Mujer": 0}
        for r in normalized_rows:
            gen = r.get("genero", "").strip().capitalize()
            if gen in gender_counts:
                gender_counts[gen] += 1
                
        total_gender = sum(gender_counts.values())
        total_gender_pct = total_gender if total_gender > 0 else 1
        genero_data = [
            { "label": "Hombre", "value": gender_counts["Hombre"], "percent": round((gender_counts["Hombre"] / total_gender_pct) * 100, 1), "color": "var(--neon-cyan)" },
            { "label": "Mujer", "value": gender_counts["Mujer"], "percent": round((gender_counts["Mujer"] / total_gender_pct) * 100, 1), "color": "var(--neon-purple)" }
        ]

        # Estadisticas de etnia
        etnia_counts = {"Blanco": 0, "Moreno": 0}
        for r in normalized_rows:
            etn = r.get("etnia", "").strip().capitalize()
            if etn in etnia_counts:
                etnia_counts[etn] += 1
                
        total_etnia = sum(etnia_counts.values())
        total_etnia_pct = total_etnia if total_etnia > 0 else 1
        etnia_data = [
            { "label": "Blanco", "value": etnia_counts["Blanco"], "percent": round((etnia_counts["Blanco"] / total_etnia_pct) * 100, 1), "color": "#f39c12" },
            { "label": "Moreno", "value": etnia_counts["Moreno"], "percent": round((etnia_counts["Moreno"] / total_etnia_pct) * 100, 1), "color": "#9b59b6" }
        ]

        # Cruce de Genero y Etnia
        cruce_counts = {
            "Hombre Blanco": 0,
            "Hombre Moreno": 0,
            "Mujer Blanco": 0,
            "Mujer Moreno": 0
        }
        for r in normalized_rows:
            gen = r.get("genero", "").strip().capitalize()
            etn = r.get("etnia", "").strip().capitalize()
            key = f"{gen} {etn}"
            if key in cruce_counts:
                cruce_counts[key] += 1

        cruce_gen_etnia = [
            { "label": "Hombre Blanco", "value": cruce_counts["Hombre Blanco"], "color": "#ff7f50" },
            { "label": "Hombre Moreno", "value": cruce_counts["Hombre Moreno"], "color": "var(--neon-cyan)" },
            { "label": "Mujer Blanco", "value": cruce_counts["Mujer Blanco"], "color": "#e0115f" },
            { "label": "Mujer Moreno", "value": cruce_counts["Mujer Moreno"], "color": "#8e44ad" }
        ]

        # Mapas solo hombres
        mapas_solo_hombres = []
        for m in unique_maps:
            s = map_stats[m]
            if "hombre" in s["genders"] and "mujer" not in s["genders"]:
                mapas_solo_hombres.append(m)
        mapas_solo_hombres.sort()

        # Estadisticas de camara
        camera_counts = {}
        for r in normalized_rows:
            cam = r.get("camara", "").strip()
            if cam:
                camera_counts[cam] = camera_counts.get(cam, 0) + 1

        camera_label_mapping = {
            "Primera Persona": "1ra Persona"
        }
        camera_colors = {
            "Realizadora": "#ff5722",
            "Libre": "#00bcd4",
            "Rail": "#9c27b0",
            "Fija": "#e91e63",
            "1ra Persona": "#4caf50",
            "Primera Persona": "#4caf50"
        }

        camara_data_list = []
        for cam, count in camera_counts.items():
            lbl = camera_label_mapping.get(cam, cam)
            color = camera_colors.get(lbl, "#7f8c8d")
            camara_data_list.append({ "label": lbl, "value": count, "color": color })
        camara_data_list.sort(key=lambda x: x["value"], reverse=True)

        # Estadisticas de duración
        duration_counts = {}
        for r in normalized_rows:
            d_val = r.get("duracion", "")
            if d_val:
                try:
                    dur = int(float(d_val))
                    duration_counts[dur] = duration_counts.get(dur, 0) + 1
                except ValueError:
                    pass

        duracion_data = []
        for dur in sorted(duration_counts.keys()):
            duracion_data.append({ "label": f"{dur}s", "value": duration_counts[dur] })

        # Action plan metrics
        mapas_solo_una_etnia = [m for m in unique_maps if len(map_stats[m]["ethnicities"]) == 1]
        fija_count = camera_counts.get("Fija", 0)

        # Plan de acción 100% dinámico
        action_plan = [
            { "id": 1, "cat": "Especie", "title": "Clasificar videos en especie", "desc": f"Clasificar los {sin_clasificar} videos marcados como 'Nan' en especie (revisión manual urgente).", "priority": "critica", "deficit": f"{sin_clasificar} videos", "status": "Pendiente", "checked": False },
            { "id": 2, "cat": "Género x Mapa", "title": "Añadir mujeres en mapas masculinos", "desc": f"Añadir personajes femeninos en los {len(mapas_solo_hombres)} mapas que solo contienen hombres (Biblioteca, Desierto, etc.).", "priority": "alta", "deficit": f"{len(mapas_solo_hombres)} mapas", "status": "Pendiente", "checked": False },
            { "id": 3, "cat": "Etnia x Mapa", "title": "Diversificar etnias en mapas exclusivos", "desc": f"Introducir variedad étnica en los {len(mapas_solo_una_etnia)} mapas que solo cuentan con un solo tipo étnico.", "priority": "alta", "deficit": f"{len(mapas_solo_una_etnia)} mapas", "status": "Pendiente", "checked": False }
        ]
        
        # Agregar tareas de cobertura dinámica para cada usuario
        act_id = 4
        for u in unique_users:
            excl_maps = exclusivas_por_usuario[u]
            if excl_maps:
                action_plan.append({
                    "id": act_id,
                    "cat": "Cobertura",
                    "title": f"Videos de otros usuarios en mapas de {u}",
                    "desc": f"Otros miembros del equipo deben grabar en los {len(excl_maps)} mapas exclusivos de {u} para completar la cobertura cruzada y evitar sesgos de captura.",
                    "priority": "media",
                    "deficit": f"{len(excl_maps)} mapas",
                    "status": "Pendiente",
                    "checked": False
                })
                act_id += 1

        action_plan.extend([
            { "id": act_id, "cat": "Cámara", "title": "Aumentar cámara Fija", "desc": f"Grabar más videos utilizando el tipo de cámara 'Fija' para nivelar el déficit (actualmente {fija_count}).", "priority": "baja", "deficit": f"{fija_count} videos", "status": "Pendiente", "checked": False },
            { "id": act_id + 1, "cat": "Cámara", "title": "Equilibrar duración del censo", "desc": "Apuntar a una mayor cantidad de grabaciones en el rango ideal de 7 a 10 segundos.", "priority": "baja", "deficit": "Varios", "status": "Pendiente", "checked": False }
        ])

        # Convertir conjuntos a listas/diccionarios para poder serializarlos a JSON
        usuarios_data = []
        for u in unique_users:
            val = videos_por_usuario[u]
            usuarios_data.append({
                "usuario": u,
                "value": val,
                "percent": round((val / total_for_pct) * 100, 1) if total_videos > 0 else 0
            })

        data = {
            "kpis": kpis,
            "especiesData": especies_data,
            "inicialMapas": inicial_mapas,
            "exclusivasMateo": exclusivas_mateo,
            "exclusivasMiguel": exclusivas_miguel,
            "exclusivasPorUsuario": exclusivas_por_usuario,
            "usuariosData": usuarios_data,
            "generoData": genero_data,
            "etniaData": etnia_data,
            "cruceGenEtnia": cruce_gen_etnia,
            "mapasSoloHombres": mapas_solo_hombres,
            "camaraData": camara_data_list,
            "duracionData": duracion_data,
            "actionPlan": action_plan
        }
        return Response(data)

class RegistroSummaryView(APIView):
    """
    Vista que analiza el Excel de Registro de Parámetros (Fase 2 — Estilizado)
    y devuelve todas las métricas de balance del dataset para el nuevo Dashboard.
    """
    permission_classes = []

    def get(self, request):
        import openpyxl
        import csv
        import re
        import os
        from django.conf import settings
        from collections import Counter

        # ──────────────────────────────────────────
        # 1. LECTURA DEL EXCEL
        # ──────────────────────────────────────────
        excel_path = os.path.join(settings.BASE_DIR, 'MATEO REGISTRO PARAMETROS.xlsx')
        if not os.path.exists(excel_path):
            return Response({'error': 'Archivo Excel no encontrado'}, status=404)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        registros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ignorar filas vacías
            if not any(v is not None for v in row[:4]):
                continue

            miembro      = str(row[0]).strip() if row[0] else ''
            id_video     = str(row[1]).strip() if row[1] is not None else ''
            productor    = str(row[2]).strip() if row[2] else ''
            estilo       = str(row[3]).strip() if row[3] else ''
            prompt_vid   = str(row[6]).strip() if row[6] else ''
            img_link     = str(row[5]).strip() if row[5] else ''
            aceptado_raw = str(row[9]).strip() if row[9] else ''
            prompt_final   = str(row[10]).strip() if len(row) > 10 and row[10] else ''
            img_arreglo    = str(row[12]).strip() if len(row) > 12 and row[12] else ''
            vid_arreglo    = str(row[14]).strip() if len(row) > 14 and row[14] else ''

            # Normalizar estado de aceptación
            aceptado_lower = aceptado_raw.lower()
            if aceptado_lower == 'si':
                estado = 'Aceptado'
            elif aceptado_lower == 'no':
                estado = 'Rechazado'
            elif aceptado_lower == 'duda':
                estado = 'Duda'
            else:
                estado = 'Sin revisar'

            # Campos derivados del pipeline
            tiene_prompt_final = bool(prompt_final and prompt_final.lower() not in ['', 'none', 'nan'])
            tiene_img_arreglo  = bool(img_arreglo and 'drive' in img_arreglo.lower())
            tiene_vid_arreglo  = bool(vid_arreglo and 'drive' in vid_arreglo.lower())
            pipeline_completo  = tiene_prompt_final and tiene_img_arreglo and tiene_vid_arreglo

            # Inferencia de movimiento de cámara desde prompt_video
            pv = prompt_vid.lower()
            if re.search(r'est[aá]tic|inm[oó]vil|fij[ao]|congelad|cero movimiento|completamente est', pv):
                movimiento_camara = 'Fija'
            elif re.search(r'tracking|sigue |seguimiento', pv):
                movimiento_camara = 'Tracking'
            elif re.search(r'orbital|gira |rotaci[oó]n|rotando|giro', pv):
                movimiento_camara = 'Orbital'
            elif re.search(r'a[eé]re|drone|elev[aá]ndose|vue?la|desde el aire', pv):
                movimiento_camara = 'Aérea'
            elif re.search(r'dolly|avanza|retrocede|zoom|se acerca|se aleja|lateral|desplaz|truck|pan ', pv):
                movimiento_camara = 'Dolly/Pan'
            else:
                movimiento_camara = 'Libre/Otro'

            # Inferencia de tipo de sujeto en movimiento
            if re.search(r'hombre|mujer|persona|personaje|corre|camina|salta|se sienta|se levanta|gesticula', pv):
                tipo_sujeto = 'Humano'
            elif re.search(r'animal|perro|gato|ping[üu]ino|tibur[oó]n|leopardo|dinosaurio|p[aá]jaro|elefante', pv):
                tipo_sujeto = 'Animal'
            elif re.search(r'coche|auto|veh[ií]culo|tren|moto|avi[oó]n|barco|cami[oó]n', pv):
                tipo_sujeto = 'Vehículo'
            elif re.search(r'agua|viento|nubes?|ondula|nieve|llama|oleaje|entorno|paisaje', pv):
                tipo_sujeto = 'Entorno'
            else:
                tipo_sujeto = 'Sin clasificar'

            registros.append({
                'miembro':           miembro,
                'id':                id_video,
                'productor':         productor,
                'estilo':            estilo,
                'estado':            estado,
                'tiene_prompt_final':  tiene_prompt_final,
                'tiene_img_arreglo':   tiene_img_arreglo,
                'tiene_vid_arreglo':   tiene_vid_arreglo,
                'pipeline_completo':   pipeline_completo,
                'movimiento_camara':   movimiento_camara,
                'tipo_sujeto':         tipo_sujeto,
                'prompt_video':        prompt_vid,
                'mapa':    '',
                'especie': '',
                'genero':  'N/A',
                'etnia':   'N/A',
            })

        total = len(registros)

        # ──────────────────────────────────────────
        # 2. JOIN CON CENSO.CSV  (mapa, especie, genero, etnia)
        # ──────────────────────────────────────────
        def norm_id(id_v):
            """Normaliza IDs: '1.0' -> '1' para el JOIN con el censo."""
            try:
                f2 = float(str(id_v).strip())
                if f2 == int(f2):
                    return str(int(f2))
            except (ValueError, TypeError):
                pass
            return str(id_v).strip()

        censo_path = os.path.join(settings.BASE_DIR, 'censo.csv')
        censo_lookup = {}
        if os.path.exists(censo_path):
            with open(censo_path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for crow in reader:
                    vid_id = norm_id(
                        crow.get('ID DE VIDEO') or crow.get('ID DE VIDEO EQUIPO') or ''
                    )
                    if vid_id:
                        censo_lookup[vid_id] = {
                            'mapa':    str(crow.get('MAPA', '')).strip(),
                            'especie': str(crow.get('ESPECIE', '')).strip(),
                            'genero':  str(crow.get('GENERO', '')).strip(),
                            'etnia':   str(crow.get('ETNIA', '')).strip(),
                        }
        for r in registros:
            c = censo_lookup.get(norm_id(r['id']), {})
            r['mapa']    = c.get('mapa', '')
            r['especie'] = c.get('especie', '')
            gen = c.get('genero', '')
            etn = c.get('etnia', '')
            r['genero']  = gen if gen.lower() not in ['nan', ''] else 'N/A'
            r['etnia']   = etn if etn.lower()  not in ['nan', ''] else 'N/A'

        # Enriquecimiento adicional desde MATEO METADATA.xlsx
        meta_excel_path2 = os.path.join(settings.BASE_DIR, 'MATEO METADATA.xlsx')
        if os.path.exists(meta_excel_path2):
            wb_m2 = openpyxl.load_workbook(meta_excel_path2, data_only=True)
            ws_m2 = wb_m2.active
            headers_m2 = [str(c.value).strip().lower() if c.value else '' for c in ws_m2[1]]
            col_mapa_m    = headers_m2.index('mapa')    if 'mapa'    in headers_m2 else None
            col_especie_m = headers_m2.index('especie') if 'especie' in headers_m2 else None
            col_genero_m  = headers_m2.index('genero')  if 'genero'  in headers_m2 else None
            col_etnia_m   = headers_m2.index('etnia')   if 'etnia'   in headers_m2 else None
            if any(c is not None for c in [col_mapa_m, col_especie_m, col_genero_m, col_etnia_m]):
                meta_enrich = {}
                for row_m in ws_m2.iter_rows(min_row=2, values_only=True):
                    if not row_m[0]:
                        continue
                    clave_m = str(row_m[0]).strip()
                    meta_enrich[clave_m] = {
                        'mapa':    str(row_m[col_mapa_m]).strip()    if col_mapa_m    is not None and row_m[col_mapa_m]    else '',
                        'especie': str(row_m[col_especie_m]).strip() if col_especie_m is not None and row_m[col_especie_m] else '',
                        'genero':  str(row_m[col_genero_m]).strip()  if col_genero_m  is not None and row_m[col_genero_m]  else '',
                        'etnia':   str(row_m[col_etnia_m]).strip()   if col_etnia_m   is not None and row_m[col_etnia_m]   else '',
                    }
                for r in registros:
                    clave_r = f"{r['id']}__{r['estilo']}__{r['miembro']}"
                    me = meta_enrich.get(clave_r, {})
                    if not r['mapa']    and me.get('mapa'):    r['mapa']    = me['mapa']
                    if not r['especie'] and me.get('especie'): r['especie'] = me['especie']
                    if r['genero'] == 'N/A' and me.get('genero'): r['genero'] = me['genero']
                    if r['etnia']  == 'N/A' and me.get('etnia'):  r['etnia']  = me['etnia']

        # ──────────────────────────────────────────
        # 3. BLOQUE 1 — DISTRIBUCIÓN GENERAL
        # ──────────────────────────────────────────
        total_pct        = total if total > 0 else 1
        estilo_counts    = Counter(r['estilo']    for r in registros if r['estilo'])
        productor_counts = Counter(r['productor'] for r in registros if r['productor'])
        miembro_counts   = Counter(r['miembro']   for r in registros if r['miembro'])
        estado_counts    = Counter(r['estado']    for r in registros)

        estilos_validos = [est for est, _ in estilo_counts.most_common() if est]

        PALETA_ESTILOS = [
            '#00f2ff',  # Anime (cian neon)
            '#f39c12',  # Cartoon (naranja)
            '#e74c3c',  # Lego (rojo)
            '#bc13fe',  # Ciberpunk (violeta)
            '#2ecc71', '#3498db', '#e91e63', '#9b59b6', '#1abc9c', '#ff5722',
        ]
        COLORES_CONOCIDOS = {
            'Anime': '#00f2ff', 'Cartoon': '#f39c12',
            'Lego':  '#e74c3c', 'Ciberpunk': '#bc13fe',
        }
        colores_estilo = {}
        _paleta_extra_idx = 4
        for _est in estilos_validos:
            if _est in COLORES_CONOCIDOS:
                colores_estilo[_est] = COLORES_CONOCIDOS[_est]
            else:
                colores_estilo[_est] = PALETA_ESTILOS[_paleta_extra_idx % len(PALETA_ESTILOS)]
                _paleta_extra_idx += 1

        estilos_data = [
            {
                'label':   est,
                'value':   estilo_counts.get(est, 0),
                'percent': round((estilo_counts.get(est, 0) / total_pct) * 100, 1),
                'color':   colores_estilo.get(est, '#7f8c8d'),
            }
            for est in estilos_validos
        ]

        PALETA_PRODUCTORES = [
            '#ff4b2b', '#00f2ff', '#bc13fe', '#2ecc71',
            '#f39c12', '#3498db', '#e91e63', '#9b59b6',
        ]
        productores_lista = sorted(productor_counts.items(), key=lambda x: -x[1])
        productores_data = [
            {
                'label':   prod,
                'value':   cnt,
                'percent': round((cnt / total_pct) * 100, 1),
                'color':   PALETA_PRODUCTORES[i % len(PALETA_PRODUCTORES)],
            }
            for i, (prod, cnt) in enumerate(productores_lista)
        ]
        color_por_productor = {
            prod: PALETA_PRODUCTORES[i % len(PALETA_PRODUCTORES)]
            for i, (prod, _) in enumerate(productores_lista)
        }

        sub_productores_data = [
            {
                'label':   mb,
                'value':   cnt,
                'percent': round((cnt / total_pct) * 100, 1),
            }
            for mb, cnt in sorted(miembro_counts.items(), key=lambda x: -x[1])
        ]

        alertas = []
        for est in estilos_validos:
            pct = round((estilo_counts.get(est, 0) / total_pct) * 100, 1)
            if pct < 15:
                alertas.append({
                    'tipo':    'critica',
                    'bloque':  'distribucion',
                    'mensaje': f'El estilo {est} tiene solo {estilo_counts.get(est, 0)} entradas ({pct}%), por debajo del mínimo recomendado del 15%.',
                })

        kpis = {
            'totalEntradas':         total,
            'totalAceptados':        estado_counts.get('Aceptado', 0),
            'totalRechazados':       estado_counts.get('Rechazado', 0),
            'totalDuda':             estado_counts.get('Duda', 0),
            'totalSinRevisar':       estado_counts.get('Sin revisar', 0),
            'tasaAceptacion':        round((estado_counts.get('Aceptado', 0) / total_pct) * 100, 1),
            'totalPipelineCompleto': sum(1 for r in registros if r['pipeline_completo']),
            'totalConPromptFinal':   sum(1 for r in registros if r['tiene_prompt_final']),
            'totalConImgArreglo':    sum(1 for r in registros if r['tiene_img_arreglo']),
            'totalConVidArreglo':    sum(1 for r in registros if r['tiene_vid_arreglo']),
        }

        # ──────────────────────────────────────────
        # 4. BLOQUE 2 — ESTADO DEL PIPELINE
        # ──────────────────────────────────────────
        aceptacion_por_estilo = []
        for est in estilos_validos:
            rows_e   = [r for r in registros if r['estilo'] == est]
            cnt_e    = len(rows_e) or 1
            acept    = sum(1 for r in rows_e if r['estado'] == 'Aceptado')
            rechaz   = sum(1 for r in rows_e if r['estado'] == 'Rechazado')
            duda     = sum(1 for r in rows_e if r['estado'] == 'Duda')
            sin_rev  = sum(1 for r in rows_e if r['estado'] == 'Sin revisar')
            pip_comp = sum(1 for r in rows_e if r['pipeline_completo'])
            tasa_a   = round((acept / cnt_e) * 100, 1)
            aceptacion_por_estilo.append({
                'estilo':         est,
                'total':          len(rows_e),
                'aceptados':      acept,
                'rechazados':     rechaz,
                'duda':           duda,
                'sinRevisar':     sin_rev,
                'tasaAceptacion': tasa_a,
                'pipelineCompleto': round((pip_comp / cnt_e) * 100, 1),
                'color':          colores_estilo.get(est, '#7f8c8d'),
            })
            if len(rows_e) > 0 and tasa_a < 50:
                alertas.append({
                    'tipo':    'advertencia',
                    'bloque':  'pipeline',
                    'mensaje': f'Tasa de aceptación del estilo {est} es {tasa_a}%, por debajo del 50% recomendado.',
                })

        if estado_counts.get('Duda', 0) > 5:
            alertas.append({
                'tipo':    'advertencia',
                'bloque':  'pipeline',
                'mensaje': f'Hay {estado_counts["Duda"]} entradas en estado "Duda" que requieren revisión manual urgente.',
            })

        aceptacion_por_miembro = []
        for mb, cnt_m in sorted(miembro_counts.items(), key=lambda x: -x[1]):
            rows_m   = [r for r in registros if r['miembro'] == mb]
            cnt_m2   = len(rows_m) or 1
            acept_m  = sum(1 for r in rows_m if r['estado'] == 'Aceptado')
            arr_m    = sum(1 for r in rows_m if r['tiene_img_arreglo'] or r['tiene_vid_arreglo'])
            aceptacion_por_miembro.append({
                'miembro':        mb,
                'total':          len(rows_m),
                'aceptados':      acept_m,
                'tasaAceptacion': round((acept_m / cnt_m2) * 100, 1),
                'tasaArreglos':   round((arr_m   / cnt_m2) * 100, 1),
            })

        # ──────────────────────────────────────────
        # 5. BLOQUE 3 — COBERTURA DE MAPAS
        # ──────────────────────────────────────────
        mapas_unicos = sorted(set(r['mapa'] for r in registros if r['mapa']))

        heatmap_estilo_mapa = {est: {} for est in estilos_validos}
        for est in estilos_validos:
            for mapa in mapas_unicos:
                heatmap_estilo_mapa[est][mapa] = sum(
                    1 for r in registros if r['estilo'] == est and r['mapa'] == mapa
                )

        mapas_stats = []
        productores_unicos = sorted(productor_counts.keys())
        for mapa in mapas_unicos:
            row_m = {'mapa': mapa, 'total': 0}
            for est in estilos_validos:
                cnt_em = heatmap_estilo_mapa[est].get(mapa, 0)
                row_m[est.lower()] = cnt_em
                row_m['total']    += cnt_em
            for prod in productores_unicos:
                row_m[prod.lower()] = sum(
                    1 for r in registros if r['mapa'] == mapa and r['productor'] == prod
                )
            mapas_stats.append(row_m)
        mapas_stats.sort(key=lambda x: -x['total'])

        mapas_gaps = []
        for mapa in mapas_unicos:
            faltantes = [est for est in estilos_validos if heatmap_estilo_mapa[est].get(mapa, 0) == 0]
            if faltantes:
                mapas_gaps.append({'mapa': mapa, 'estilosFaltantes': faltantes})

        if mapas_gaps:
            alertas.append({
                'tipo':    'advertencia',
                'bloque':  'mapas',
                'mensaje': f'{len(mapas_gaps)} mapas tienen 0 videos en al menos un estilo. Revisar gaps de cobertura.',
            })

        # ──────────────────────────────────────────
        # 6. BLOQUE 4 — BALANCE DE ESPECIE
        # ──────────────────────────────────────────
        especie_por_estilo = []
        for est in estilos_validos:
            rows_e  = [r for r in registros if r['estilo'] == est]
            cnt_e   = len(rows_e) or 1
            humano  = sum(1 for r in rows_e if r['especie'].lower() == 'humano')
            animal  = sum(1 for r in rows_e if r['especie'].lower() == 'animal')
            entorno = cnt_e - humano - animal
            pct_an  = round((animal  / cnt_e) * 100, 1)
            pct_en  = round((entorno / cnt_e) * 100, 1)
            especie_por_estilo.append({
                'estilo':     est,
                'total':      len(rows_e),
                'humano':     humano,
                'animal':     animal,
                'entorno':    entorno,
                'pctHumano':  round((humano  / cnt_e) * 100, 1),
                'pctAnimal':  pct_an,
                'pctEntorno': pct_en,
                'color':      colores_estilo.get(est, '#7f8c8d'),
            })
            if len(rows_e) > 0:
                if pct_an < 20:
                    alertas.append({'tipo': 'advertencia', 'bloque': 'especie',
                                    'mensaje': f'{est}: solo {pct_an}% de animales (mínimo recomendado: 20%).'})
                if pct_en > 40:
                    alertas.append({'tipo': 'advertencia', 'bloque': 'especie',
                                    'mensaje': f'{est}: {pct_en}% sin sujeto clasificado (máximo recomendado: 40%).'})

        # ──────────────────────────────────────────
        # 7. BLOQUE 5 — GÉNERO Y ETNIA (solo humanos)
        # ──────────────────────────────────────────
        genero_por_estilo = []
        for est in estilos_validos:
            rows_h   = [r for r in registros if r['estilo'] == est and r['especie'].lower() == 'humano']
            cnt_h    = len(rows_h) or 1
            hombres  = sum(1 for r in rows_h if r['genero'].lower() == 'hombre')
            mujeres  = sum(1 for r in rows_h if r['genero'].lower() == 'mujer')
            blanco   = sum(1 for r in rows_h if r['etnia'].lower()  == 'blanco')
            moreno   = sum(1 for r in rows_h if r['etnia'].lower()  == 'moreno')
            pct_h    = round((hombres / cnt_h) * 100, 1)
            pct_b    = round((blanco  / cnt_h) * 100, 1)
            genero_por_estilo.append({
                'estilo':       est,
                'totalHumanos': len(rows_h),
                'hombres':      hombres,
                'mujeres':      mujeres,
                'blanco':       blanco,
                'moreno':       moreno,
                'pctHombres':   pct_h,
                'pctMujeres':   round((mujeres / cnt_h) * 100, 1),
                'pctBlanco':    pct_b,
                'pctMoreno':    round((moreno  / cnt_h) * 100, 1),
                'color':        colores_estilo.get(est, '#7f8c8d'),
            })
            if len(rows_h) > 0:
                if pct_h > 60:
                    alertas.append({'tipo': 'advertencia', 'bloque': 'diversidad',
                                    'mensaje': f'{est}: {pct_h}% hombres, supera el umbral 60/40 de balance de género.'})
                if pct_b > 65:
                    alertas.append({'tipo': 'advertencia', 'bloque': 'diversidad',
                                    'mensaje': f'{est}: {pct_b}% blancos, supera el umbral 65/35 de balance étnico.'})

        # ──────────────────────────────────────────
        # 8. BLOQUE 6 — DIVERSIDAD TÉCNICA
        # ──────────────────────────────────────────
        colores_camara = {
            'Fija':       '#e91e63', 'Tracking':  '#00bcd4',
            'Dolly/Pan':  '#ff9800', 'Orbital':   '#9c27b0',
            'Aérea':      '#4caf50', 'Libre/Otro':'#607d8b',
        }
        colores_sujeto = {
            'Humano':     '#00f2ff', 'Animal':    '#2ecc71',
            'Vehículo':   '#f39c12', 'Entorno':   '#9b59b6',
            'Sin clasificar': '#7f8c8d',
        }

        tecnico_por_estilo = []
        cam_global = Counter(r['movimiento_camara'] for r in registros)
        for est in estilos_validos:
            rows_e   = [r for r in registros if r['estilo'] == est]
            cnt_e    = len(rows_e) or 1
            cam_cnt  = Counter(r['movimiento_camara'] for r in rows_e)
            suj_cnt  = Counter(r['tipo_sujeto']       for r in rows_e)
            pct_fija = round((cam_cnt.get('Fija', 0) / cnt_e) * 100, 1)
            tecnico_por_estilo.append({
                'estilo':  est,
                'total':   len(rows_e),
                'pctFija': pct_fija,
                'camaraData': [
                    {'label': cam, 'value': cnt, 'percent': round((cnt / cnt_e) * 100, 1),
                     'color': colores_camara.get(cam, '#7f8c8d')}
                    for cam, cnt in sorted(cam_cnt.items(), key=lambda x: -x[1])
                ],
                'sujetoData': [
                    {'label': suj, 'value': cnt, 'percent': round((cnt / cnt_e) * 100, 1),
                     'color': colores_sujeto.get(suj, '#7f8c8d')}
                    for suj, cnt in sorted(suj_cnt.items(), key=lambda x: -x[1])
                ],
                'color': colores_estilo.get(est, '#7f8c8d'),
            })
            if len(rows_e) > 0 and pct_fija > 60:
                alertas.append({'tipo': 'advertencia', 'bloque': 'tecnico',
                                'mensaje': f'{est}: {pct_fija}% de cámara fija. Supera el 60%, poca variedad de movimiento.'})

        total_pct = sum(cam_global.values()) or 1
        camara_global_data = [
            {'label': cam, 'value': cnt, 'percent': round((cnt / total_pct) * 100, 1),
             'color': colores_camara.get(cam, '#7f8c8d')}
            for cam, cnt in sorted(cam_global.items(), key=lambda x: -x[1])
        ]

        # ─── BLOQUE 7: CALIDAD DEL PIPELINE DE ARREGLOS ───────────────
        arreglos_por_estilo = []
        for est in estilos_validos:
            rows_e    = [r for r in registros if r['estilo'] == est]
            cnt_e     = len(rows_e) or 1
            con_img   = sum(1 for r in rows_e if r['tiene_img_arreglo'])
            con_vid   = sum(1 for r in rows_e if r['tiene_vid_arreglo'])
            con_pf    = sum(1 for r in rows_e if r['tiene_prompt_final'])
            con_arr   = [r for r in rows_e if r['tiene_img_arreglo'] or r['tiene_vid_arreglo']]
            arr_acept = sum(1 for r in con_arr if r['estado'] == 'Aceptado')
            arr_cnt   = len(con_arr) or 1
            arreglos_por_estilo.append({
                'estilo':            est,
                'total':             len(rows_e),
                'conImgArreglo':     con_img,
                'conVidArreglo':     con_vid,
                'conPromptFinal':    con_pf,
                'pctImgArreglo':     round((con_img / cnt_e) * 100, 1),
                'pctVidArreglo':     round((con_vid / cnt_e) * 100, 1),
                'pctPromptFinal':    round((con_pf  / cnt_e) * 100, 1),
                'arregloAceptacion': round((arr_acept / arr_cnt) * 100, 1),
                'color':             colores_estilo.get(est, '#7f8c8d'),
            })

        top_arreglos_miembro = []
        for mb in sorted(miembro_counts.keys()):
            rows_m     = [r for r in registros if r['miembro'] == mb]
            cnt_m2     = len(rows_m) or 1
            arreglos_m = sum(1 for r in rows_m if r['tiene_img_arreglo'] or r['tiene_vid_arreglo'])
            top_arreglos_miembro.append({
                'miembro':      mb,
                'total':        len(rows_m),
                'conArreglos':  arreglos_m,
                'tasaArreglos': round((arreglos_m / cnt_m2) * 100, 1),
            })
        top_arreglos_miembro.sort(key=lambda x: -x['tasaArreglos'])

        # ─── BLOQUE 10: REGISTROS CRUDOS + META DE DURACIÓN ──────────
        meta_excel_path = os.path.join(settings.BASE_DIR, 'MATEO METADATA.xlsx')
        meta_por_clave  = {}

        if os.path.exists(meta_excel_path):
            wb_meta = openpyxl.load_workbook(meta_excel_path, data_only=True)
            ws_meta = wb_meta.active
            for row_m in ws_meta.iter_rows(min_row=2, values_only=True):
                if not row_m[0]:
                    continue
                clave_m = str(row_m[0]).strip()
                meta_por_clave[clave_m] = {
                    'dur_original_s':   row_m[4],
                    'dur_estilizado_s': row_m[5],
                    'fps':              row_m[6],
                    'resolucion':       row_m[7],
                    'size_mb':          row_m[8],
                }

        registros_para_frontend = []
        alertas_duracion = []
        for r in registros:
            clave_r   = f"{r['id']}__{r['estilo']}__{r['miembro']}"
            meta_r    = meta_por_clave.get(clave_r, {})
            dur_orig  = meta_r.get('dur_original_s')
            dur_estil = meta_r.get('dur_estilizado_s')
            alerta_dur = False
            if dur_orig and dur_estil and (dur_estil / dur_orig) < 0.8:
                alerta_dur = True
                pct_r = round((dur_estil / dur_orig) * 100, 1)
                alertas_duracion.append({
                    'tipo':   'advertencia',
                    'bloque': 'duracion',
                    'mensaje': (f'ID {r["id"]} ({r["estilo"]}): video estilizado dura '
                                f'{dur_estil}s vs {dur_orig}s original '
                                f'({pct_r}% retención). Posible corte de video.'),
                    'accion': None,
                })
            registros_para_frontend.append({
                'id':               r['id'],
                'miembro':          r['miembro'],
                'productor':        r['productor'],
                'estilo':           r['estilo'],
                'estado':           r['estado'],
                'mapa':             r['mapa'],
                'especie':          r['especie'],
                'genero':           r['genero'],
                'etnia':            r['etnia'],
                'tieneArreglo':     r['tiene_img_arreglo'] or r['tiene_vid_arreglo'],
                'movimientoCamara': r['movimiento_camara'],
                'tipoSujeto':       r['tipo_sujeto'],
                'duracionOriginal':  dur_orig,
                'duracionEstilizado': dur_estil,
                'alertaDuracion':   alerta_dur,
            })

        alertas.extend(alertas_duracion)
        if alertas_duracion:
            alertas.append({'tipo': 'informativa', 'bloque': 'duracion',
                            'mensaje': f'{len(alertas_duracion)} video(s) con retención < 80% detectados.',
                            'accion': None})

        duracion_por_estilo = []
        for est in estilos_validos:
            rows_con = [r for r in registros_para_frontend
                        if r['estilo'] == est
                        and r['duracionOriginal'] is not None
                        and r['duracionEstilizado'] is not None]
            if not rows_con:
                duracion_por_estilo.append({'estilo': est, 'conDatos': 0,
                    'avgDurOriginal': None, 'avgDurEstilizado': None,
                    'pctRetencion': None, 'videosConAlerta': 0,
                    'color': colores_estilo.get(est, '#7f8c8d')})
                continue
            avg_o = round(sum(r['duracionOriginal']   for r in rows_con) / len(rows_con), 2)
            avg_e = round(sum(r['duracionEstilizado'] for r in rows_con) / len(rows_con), 2)
            duracion_por_estilo.append({
                'estilo': est, 'conDatos': len(rows_con),
                'avgDurOriginal': avg_o, 'avgDurEstilizado': avg_e,
                'pctRetencion': round((avg_e / avg_o) * 100, 1) if avg_o > 0 else None,
                'videosConAlerta': sum(1 for r in rows_con if r['alertaDuracion']),
                'color': colores_estilo.get(est, '#7f8c8d'),
            })

        # ─── BLOQUE 11: ESTADÍSTICAS DE METADATA ──────────
        metadata_stats = {
            'totalRegistros': 0, 'conMetadata': 0, 'sinMetadata': 0,
            'pctCobertura': 0,
            'fpsPorEstilo': [],
            'resPorEstilo': [],
            'tamanioPorEstilo': [],
            'distribucionFps': [],
            'distribucionDurOrig': [],
        }

        if os.path.exists(meta_excel_path):
            wb_ms = openpyxl.load_workbook(meta_excel_path, data_only=True)
            ws_ms = wb_ms.active
            headers_ms = [str(c.value).strip().lower() if c.value else '' for c in ws_ms[1]]

            meta_rows = []
            for rm in ws_ms.iter_rows(min_row=2, values_only=True):
                if not rm[0]:
                    continue
                row_dict = {headers_ms[i]: rm[i] for i in range(min(len(headers_ms), len(rm)))}
                meta_rows.append(row_dict)

            total_meta   = len(meta_rows)
            con_meta     = sum(1 for r in meta_rows if r.get('meta_ok'))
            sin_meta     = total_meta - con_meta
            pct_cobertura = round((con_meta / total_meta) * 100, 1) if total_meta > 0 else 0

            fps_por_estilo = []
            for est in estilos_validos:
                fps_vals = [
                    float(r['fps_estilizado']) for r in meta_rows
                    if r.get('estilo') == est and r.get('fps_estilizado')
                ]
                if fps_vals:
                    fps_por_estilo.append({
                        'estilo': est,
                        'avgFps': round(sum(fps_vals) / len(fps_vals), 1),
                        'conDatos': len(fps_vals),
                        'color': colores_estilo.get(est, '#7f8c8d'),
                    })

            res_por_estilo = []
            for est in estilos_validos:
                res_vals = [
                    str(r['resolucion_estilizado']) for r in meta_rows
                    if r.get('estilo') == est and r.get('resolucion_estilizado')
                ]
                if res_vals:
                    conteos = Counter(res_vals)
                    mas_comun = conteos.most_common(1)[0][0]
                    res_por_estilo.append({
                        'estilo': est,
                        'resolucionMasComun': mas_comun,
                        'conDatos': len(res_vals),
                        'conteos': [{'res': r, 'cnt': c} for r, c in conteos.most_common(5)],
                        'color': colores_estilo.get(est, '#7f8c8d'),
                    })

            tamanio_por_estilo = []
            for est in estilos_validos:
                size_vals = [
                    float(r['size_mb_estilizado']) for r in meta_rows
                    if r.get('estilo') == est and r.get('size_mb_estilizado')
                ]
                if size_vals:
                    tamanio_por_estilo.append({
                        'estilo': est,
                        'avgSizeMb': round(sum(size_vals) / len(size_vals), 1),
                        'totalSizeMb': round(sum(size_vals), 1),
                        'conDatos': len(size_vals),
                        'color': colores_estilo.get(est, '#7f8c8d'),
                    })

            fps_global = [
                str(int(float(r['fps_estilizado']))) for r in meta_rows
                if r.get('fps_estilizado')
            ]
            fps_dist = [
                {'fps': k, 'cantidad': v}
                for k, v in sorted(Counter(fps_global).items(), key=lambda x: -x[1])
            ]

            rangos_dur = {'<5s': 0, '5-10s': 0, '10-20s': 0, '>20s': 0}
            for r in meta_rows:
                d = r.get('dur_original_s')
                if d is not None:
                    d = float(d)
                    if d < 5:   rangos_dur['<5s']   += 1
                    elif d < 10: rangos_dur['5-10s'] += 1
                    elif d < 20: rangos_dur['10-20s']+= 1
                    else:        rangos_dur['>20s']  += 1
            dur_dist = [{'rango': k, 'cantidad': v} for k, v in rangos_dur.items()]

            metadata_stats = {
                'totalRegistros':    total_meta,
                'conMetadata':       con_meta,
                'sinMetadata':       sin_meta,
                'pctCobertura':      pct_cobertura,
                'fpsPorEstilo':      fps_por_estilo,
                'resPorEstilo':      res_por_estilo,
                'tamanioPorEstilo':  tamanio_por_estilo,
                'distribucionFps':   fps_dist,
                'distribucionDurOrig': dur_dist,
            }

        return Response({
            'kpis':                   kpis,
            'estilosData':            estilos_data,
            'productoresData':        productores_data,
            'productoresUnicos':      [p for p, _ in productores_lista],
            'coloresPorProductor':    color_por_productor,
            'subProductoresData':     sub_productores_data,
            'aceptacionPorEstilo':    aceptacion_por_estilo,
            'aceptacionPorMiembro':   aceptacion_por_miembro,
            'heatmapEstiloMapa':      heatmap_estilo_mapa,
            'mapasStats':             mapas_stats,
            'mapasGaps':              mapas_gaps,
            'especiePorEstilo':       especie_por_estilo,
            'generoPorEstilo':        genero_por_estilo,
            'tecnicoPorEstilo':       tecnico_por_estilo,
            'camaraGlobal':           camara_global_data,
            'arreglosPorEstilo':      arreglos_por_estilo,
            'topArreglosMiembro':     top_arreglos_miembro[:5],
            'alertas':                alertas,
            'estilosValidos':         estilos_validos,
            'coloresEstilo':          colores_estilo,
            'registrosCrudos':        registros_para_frontend,
            'duracionPorEstilo':      duracion_por_estilo,
            'metadataStats':          metadata_stats,
        })


class SyncFromSheetsView(APIView):
    permission_classes = []

    def post(self, request):
        import requests as req_lib
        import urllib3
        import re
        import io
        import os
        import openpyxl
        from datetime import datetime
        from django.conf import settings

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        SHEETS_URL = (
            'https://docs.google.com/spreadsheets/d/'
            '1Ga5zMekIlVjHKhxkfGBIhAiCh0H3zGYf0TOoMkoctj4'
            '/export?format=xlsx&gid=2173056'
        )
        NUM_COLS  = 15
        COLS_LINK = {5, 7, 8, 12, 14}
        excel_path = os.path.join(settings.BASE_DIR, 'MATEO REGISTRO PARAMETROS.xlsx')

        def extraer_celda(cell, col_idx=None):
            val = cell.value
            texto = str(val).strip() if val is not None else ''
            if texto.lower() == 'none':
                texto = ''

            if col_idx in COLS_LINK:
                if cell.hyperlink and cell.hyperlink.target:
                    return str(cell.hyperlink.target).strip()
                m = re.match(r'=HYPERLINK\("([^"]+)"', texto, re.IGNORECASE)
                if m:
                    return m.group(1)
                if texto.lower().startswith('http'):
                    return texto

            return texto

        def hacer_clave(vals):
            return f"{str(vals[1]).strip()}__{str(vals[3]).strip()}__{str(vals[0]).strip()}"

        try:
            resp = req_lib.get(SHEETS_URL, timeout=60, verify=False)
            resp.raise_for_status()
        except Exception as e:
            return Response({'error': f'No se pudo descargar el Sheets: {e}'}, status=502)

        try:
            wb_sheets = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
            ws_sheets = wb_sheets.active
        except Exception as e:
            return Response({'error': f'No se pudo leer el XLSX: {e}'}, status=500)

        sheets_por_clave = {}
        for row in ws_sheets.iter_rows(min_row=2):
            vals = [
                extraer_celda(row[i], col_idx=i) if i < len(row) else ''
                for i in range(NUM_COLS)
            ]
            if not any(vals[:4]) or not vals[1].strip():
                continue
            sheets_por_clave[hacer_clave(vals)] = vals

        if not os.path.exists(excel_path):
            return Response({'error': 'Excel local no encontrado'}, status=404)

        wb_local = openpyxl.load_workbook(excel_path)
        ws_local = wb_local.active

        excel_por_clave = {}
        fila_maxima = 1
        for idx, row in enumerate(ws_local.iter_rows(min_row=2), start=2):
            if len(row) < 4:
                continue
            vals_loc = [
                extraer_celda(row[i], col_idx=i) if i < len(row) else ''
                for i in range(NUM_COLS)
            ]
            if not any(vals_loc[:4]) or not vals_loc[1].strip():
                continue
            excel_por_clave[hacer_clave(vals_loc)] = idx
            fila_maxima = idx

        nuevas      = 0
        actualizadas = 0

        for clave, fila_s in sheets_por_clave.items():
            if clave not in excel_por_clave:
                fila_maxima += 1
                for col_i, val in enumerate(fila_s, 1):
                    ws_local.cell(row=fila_maxima, column=col_i, value=val)
                nuevas += 1
            else:
                fila_idx   = excel_por_clave[clave]
                fila_excel = [
                    extraer_celda(ws_local.cell(fila_idx, c + 1), col_idx=c)
                    for c in range(NUM_COLS)
                ]
                if fila_excel != fila_s:
                    for col_i, val in enumerate(fila_s, 1):
                        ws_local.cell(row=fila_idx, column=col_i, value=val)
                    actualizadas += 1

        if nuevas + actualizadas > 0:
            wb_local.save(excel_path)

        return Response({
            'ok':           True,
            'nuevas':       nuevas,
            'actualizadas': actualizadas,
            'totalSheets':  len(sheets_por_clave),
            'totalExcel':   len(excel_por_clave),
            'timestamp':    datetime.now().isoformat(),
        })

class ExtractMetadataView(APIView):
    permission_classes = []

    def get(self, request):
        import os, openpyxl
        from django.conf import settings
        meta_path = os.path.join(settings.BASE_DIR, 'MATEO METADATA.xlsx')
        if not os.path.exists(meta_path):
            return Response({
                'ok':             False,
                'mensaje':        'No hay Excel de metadata todavia. Ejecuta: python manage.py extract_metadata',
                'totalRegistros': 0,
                'registrosOk':    0,
                'lastUpdated':    None,
            })
        wb = openpyxl.load_workbook(meta_path, data_only=True)
        ws = wb.active
        total = ok_count = 0
        last_updated = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            total += 1
            if row[12]:                       
                ok_count += 1
            if row[13]:                       
                ts = str(row[13])
                if last_updated is None or ts > last_updated:
                    last_updated = ts
        return Response({
            'ok':             True,
            'totalRegistros': total,
            'registrosOk':    ok_count,
            'registrosFail':  total - ok_count,
            'lastUpdated':    last_updated,
            'excelPath':      meta_path,
        })

    def post(self, request):
        import threading
        import subprocess
        import sys
        import os
        from django.conf import settings

        limite = int(request.data.get('limit', 0))
        manage_py = os.path.join(settings.BASE_DIR, 'manage.py')
        python = sys.executable

        cmd = [python, manage_py, 'extract_metadata']
        if limite > 0:
            cmd += ['--limit', str(limite)]

        def ejecutar():
            subprocess.run(cmd, cwd=settings.BASE_DIR)

        hilo = threading.Thread(target=ejecutar, daemon=True)
        hilo.start()

        return Response({
            'ok': True,
            'mensaje': f'Extracción iniciada en segundo plano (limit={limite or "sin límite"}).',
            'nota': 'Consulta GET /api/sheets/extract-metadata/ para ver el progreso.',
        })
