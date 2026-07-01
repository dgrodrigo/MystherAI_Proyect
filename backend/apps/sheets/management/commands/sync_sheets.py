import csv, os, io, re
import requests as req
import openpyxl
from django.core.management.base import BaseCommand
from django.db import IntegrityError
from apps.sheets.models import VideoMetadata

SHEETS_ID = '1Ga5zMekIlVjHKhxkfGBIhAiCh0H3zGYf0TOoMkoctj4'
XLSX_URL  = f'https://docs.google.com/spreadsheets/d/{SHEETS_ID}/export?format=xlsx'

# Columnas de Registro que contienen URLs (índice 0-based)
# 5=imagen, 7=video, 8=video orginal, 12=Imagen arreglo, 14=Video arreglo
REGISTRO_LINK_COLS = {5, 7, 8, 12, 14}

class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        VideoMetadata.objects.all().delete()
        self.stdout.write("Limpiando y Re-importando datos...")

        def safe_create(obj):
            try:
                obj.save()
            except IntegrityError:
                pass

        # ---------------------------------------------------------------
        # REGISTRO — leer desde XLSX para preservar hipervínculos de Smart Chips
        # ---------------------------------------------------------------
        count = 0
        try:
            self.stdout.write("Descargando Registro desde Google Sheets (XLSX)...")
            resp = req.get(XLSX_URL, timeout=60)
            resp.raise_for_status()
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)

            # Buscar la hoja de Registro (por nombre o la primera disponible)
            sheet_names = [s.lower() for s in wb.sheetnames]
            if 'registro' in sheet_names:
                ws = wb[wb.sheetnames[sheet_names.index('registro')]]
            else:
                ws = wb.active
            self.stdout.write(f"  Hoja: '{ws.title}' ({ws.max_row} filas)")

            def extraer_celda(cell, col_idx=None):
                """Extrae el valor de la celda; para columnas de URL usa el hipervínculo."""
                val   = cell.value
                texto = str(val).strip() if val is not None else ''
                if texto.lower() in ('none', ''):
                    texto = ''

                if col_idx in REGISTRO_LINK_COLS:
                    # 1) hipervínculo explícito (incluye Smart Chips exportados)
                    if cell.hyperlink and cell.hyperlink.target:
                        return str(cell.hyperlink.target).strip()
                    # 2) fórmula =HYPERLINK("url", ...)
                    m = re.match(r'=HYPERLINK\("([^"]+)"', texto, re.IGNORECASE)
                    if m:
                        return m.group(1)
                    # 3) texto directo que ya es una URL
                    if texto.lower().startswith('http'):
                        return texto
                    # 4) Smart Chip exportado como nombre de archivo → no tenemos URL
                    return ''

                return texto

            # Leer encabezados de la primera fila
            headers = [extraer_celda(c) for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col_idx(name):
                """Devuelve el índice de columna para un nombre de encabezado."""
                name_l = name.lower().strip()
                for i, h in enumerate(headers):
                    if h.lower().strip() == name_l:
                        return i
                return None

            idx_id    = col_idx('id')
            idx_mb    = col_idx('miembro')
            idx_mm    = col_idx('mateo/miguel')
            idx_est   = col_idx('estilizado')
            idx_pi    = col_idx('prompt imagen')
            idx_img   = col_idx('imagen')
            idx_pv    = col_idx('prompt video')
            idx_vid   = col_idx('video')
            idx_orig  = col_idx('video orginal') or col_idx('video original')
            idx_acep  = col_idx('aceptado')
            idx_pf    = col_idx('prompt final')

            def cell_val(row_cells, idx, is_link=False):
                if idx is None or idx >= len(row_cells):
                    return ''
                return extraer_celda(row_cells[idx],
                                     col_idx=idx if is_link else None)

            skipped = 0
            for row in ws.iter_rows(min_row=2):
                row_cells = list(row)
                v_id      = cell_val(row_cells, idx_id)
                video_url = cell_val(row_cells, idx_vid, is_link=True)
                img_url   = cell_val(row_cells, idx_img, is_link=True)
                prompt    = cell_val(row_cells, idx_pv)

                # Saltar filas incompletas: sin ID, sin video, sin imagen o sin prompt
                if not v_id or not video_url or not img_url or not prompt:
                    skipped += 1
                    continue

                obj = VideoMetadata(
                    video_id    = v_id,
                    tipo        = 'registro',
                    usuario     = cell_val(row_cells, idx_mb),
                    mateo_miguel= cell_val(row_cells, idx_mm),
                    estilizado  = cell_val(row_cells, idx_est),
                    prompt_imagen = cell_val(row_cells, idx_pi),
                    imagen_link = img_url,
                    prompt_video= prompt,
                    drive_link  = video_url,
                    video_original_link = cell_val(row_cells, idx_orig, is_link=True),
                    aceptado    = cell_val(row_cells, idx_acep),
                    prompt_final= cell_val(row_cells, idx_pf),
                )
                safe_create(obj)
                count += 1

            self.stdout.write(f"✅ Registro importado desde Sheets (XLSX): {count} filas ({skipped} incompletas omitidas).")

        except Exception as e:
            self.stdout.write(f"⚠️  XLSX no disponible ({e}), usando registro.csv local...")
            registro_path = 'registro.csv'
            if os.path.exists(registro_path):
                def get_v(row, *keys):
                    for k in keys:
                        for rk, rv in row.items():
                            if rk and rk.lower().strip() == k.lower().strip():
                                v = rv.strip()
                                if v and not v.startswith('http') and '.' in v and '/' not in v:
                                    return ''
                                return v
                    return ''
                with open(registro_path, mode='r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        v_id = get_v(row, 'Id')
                        if not v_id:
                            continue
                        obj = VideoMetadata(
                            video_id=v_id, tipo='registro',
                            usuario=get_v(row, 'Miembro'),
                            mateo_miguel=get_v(row, 'Mateo/Miguel'),
                            estilizado=get_v(row, 'Estilizado'),
                            prompt_imagen=get_v(row, 'Prompt imagen'),
                            imagen_link=get_v(row, 'imagen'),
                            prompt_video=get_v(row, 'prompt video'),
                            drive_link=get_v(row, 'video'),
                            video_original_link=get_v(row, 'video orginal', 'video original'),
                            aceptado=get_v(row, 'ACEPTADO'),
                            prompt_final=get_v(row, 'Prompt Final'),
                        )
                        safe_create(obj)
                        count += 1
                self.stdout.write(f"✅ Registro importado desde CSV: {count} filas.")
            else:
                self.stdout.write("❌ No se encontró registro.csv")

        # ---------------------------------------------------------------
        # CENSO — CSV local (no tiene Smart Chips)
        # ---------------------------------------------------------------
        censo_path = 'censo.csv'
        if os.path.exists(censo_path):
            count = 0
            with open(censo_path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    v_id = row.get('ID DE VIDEO', '').strip()
                    if not v_id:
                        continue
                    obj = VideoMetadata(
                        video_id=v_id, tipo='censo',
                        usuario=row.get('usuario', '').strip(),
                        id_video_equipo=row.get('ID DE VIDEO EQUIPO', '').strip(),
                        drive_link=row.get('LINK', '').strip(),
                        mapa=row.get('MAPA', '').strip(),
                        genero=row.get('GENERO', '').strip(),
                        etnia=row.get('ETNIA', '').strip(),
                        duracion=row.get('DURACION', '').strip(),
                        camara=row.get('CAMARA', '').strip(),
                        especie=row.get('ESPECIE', '').strip(),
                    )
                    safe_create(obj)
                    count += 1
            self.stdout.write(f"✅ Censo importado: {count} filas.")
        else:
            self.stdout.write("❌ No se encontró censo.csv")
