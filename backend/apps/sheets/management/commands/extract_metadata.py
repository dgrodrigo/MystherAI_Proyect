"""
Comando Django: python manage.py extract_metadata
Extrae metadatos (duración, resolución, fps) de videos e imágenes en Google Drive
usando yt-dlp para videos y PIL para imágenes.

Guarda los resultados en MATEO METADATA.xlsx (mismo directorio que
MATEO REGISTRO PARAMETROS.xlsx) usando la clave compuesta ID__Estilo__Miembro
para relacionar ambos excels de forma directa.

Uso:
  python manage.py extract_metadata              # procesa los pendientes
  python manage.py extract_metadata --limit 5    # solo 5 registros (para prueba)
  python manage.py extract_metadata --id 62      # solo los registros con ID=62
  python manage.py extract_metadata --force      # re-extraer aunque ya tengan datos
  python manage.py extract_metadata --reset      # borrar Excel de metadata y empezar de cero
"""
import os
import re
import csv
import subprocess
import sys
from datetime import datetime

from django.conf import settings
from django.core.management.base import BaseCommand
import openpyxl
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


REGISTRO_EXCEL = 'MATEO REGISTRO PARAMETROS.xlsx'
METADATA_EXCEL = 'MATEO METADATA.xlsx'
CENSO_CSV      = 'censo.csv'   # contiene mapa, especie, genero, etnia del video original

# Regex para extraer el file_id de un link de Google Drive
DRIVE_FILE_ID_RE = re.compile(r'drive\.google\.com/file/d/([A-Za-z0-9_\-]+)')

# Indices (0-based) de columnas de links en MATEO REGISTRO PARAMETROS.xlsx
# F=5 (imagen), H=7 (video estilizado), I=8 (video original)
COLS_LINK = {5, 7, 8, 12, 14}

# Columnas del Excel de metadata (en orden, para los encabezados)
ENCABEZADOS = [
    'clave',              # A: ID__Estilo__Miembro (clave de relación)
    'id_registro',        # B: ID numérico del video
    'estilo',             # C: Estilo visual
    'miembro',            # D: Sub-productor/Miembro
    'dur_original_s',     # E: Duración video original (segundos)
    'dur_estilizado_s',   # F: Duración video estilizado (segundos)
    'fps_estilizado',     # G: FPS del video estilizado
    'resolucion_estilizado',  # H: Resolución ej: 1920x1080
    'size_mb_estilizado', # I: Tamaño en MB del video estilizado
    'img_width',          # J: Ancho imagen generada (px)
    'img_height',         # K: Alto imagen generada (px)
    'img_size_kb',        # L: Tamaño imagen (KB)
    'meta_ok',            # M: True si se pudo extraer al menos un dato
    'ultima_extraccion',  # N: Timestamp ISO de la última extracción
    'mapa',               # O: Localización/escenario del video
    'especie',            # P: Humano / Animal / Sin clasificar
    'genero',             # Q: Hombre / Mujer / N/A
    'etnia',              # R: Blanco / Moreno / N/A
]


# ── Utilidades ─────────────────────────────────────────────────────────────

def extraer_celda(cell, col_idx=None):
    """
    Lee el valor de una celda respetando el tipo de columna:
    - Columnas de LINK (col_idx en COLS_LINK): prioriza hipervinculo/formula HYPERLINK/URL.
    - Columnas de TEXTO: retorna siempre el valor de texto de la celda.
    """
    val = cell.value
    texto = str(val).strip() if val is not None else ''
    if texto.lower() == 'none':
        texto = ''
    if col_idx in COLS_LINK:
        if cell.hyperlink and cell.hyperlink.target:
            return str(cell.hyperlink.target).strip()
        m = re.match(r'=HYPERLINK\("([^"]+)"', texto, re.IGNORECASE)
        if m:
            return m.group(1)
        if texto.lower().startswith('http'):
            return texto
    return texto


def extraer_file_id(url):
    """Extrae el file_id de una URL de Google Drive."""
    if not url:
        return None
    m = DRIVE_FILE_ID_RE.search(str(url))
    return m.group(1) if m else None


def extraer_metadata_video(file_id, timeout=30):
    """
    Usa yt-dlp para extraer metadatos del video sin descargarlo.
    Retorna dict con duration_s, resolution, fps, size_mb o None si falla.
    """
    url = f'https://drive.google.com/file/d/{file_id}/view'
    try:
        resultado = subprocess.run(
            [
                sys.executable, '-m', 'yt_dlp',
                '--dump-json', '--no-playlist',
                '--socket-timeout', '15',
                url
            ],
            capture_output=True, text=True, timeout=timeout
        )
        if resultado.returncode != 0:
            return None
        import json
        info = json.loads(resultado.stdout)
        duracion = info.get('duration')
        width    = info.get('width')
        height   = info.get('height')
        fps      = info.get('fps')
        filesize = info.get('filesize') or info.get('filesize_approx')
        # Buscar la mejor resolución entre los formatos disponibles
        if not width and 'formats' in info:
            for fmt in reversed(info['formats']):
                if fmt.get('width') and fmt.get('height'):
                    width  = fmt['width']
                    height = fmt['height']
                    fps    = fps or fmt.get('fps')
                    break
        return {
            'duration_s': round(float(duracion), 2) if duracion else None,
            'resolution': f'{width}x{height}' if width and height else None,
            'fps':        round(float(fps), 2) if fps else None,
            'size_mb':    round(filesize / 1024 / 1024, 2) if filesize else None,
            'ok':         True,
        }
    except (subprocess.TimeoutExpired, Exception):
        return None


def extraer_metadata_imagen(file_id, timeout=20):
    """
    Descarga (parcialmente) la imagen para obtener dimensiones con PIL.
    Retorna dict con width, height, size_kb o None si falla.
    """
    try:
        import requests
        url  = f'https://drive.google.com/uc?export=download&id={file_id}'
        resp = requests.get(url, stream=True, timeout=timeout,
                            allow_redirects=True, verify=False)
        resp.raise_for_status()
        size_bytes = int(resp.headers.get('Content-Length', 0))
        width, height = None, None
        try:
            from PIL import Image
            import io
            contenido = b''
            for chunk in resp.iter_content(65536):
                contenido += chunk
                if len(contenido) > 500 * 1024:   # solo primeros 500 KB
                    break
            img = Image.open(io.BytesIO(contenido))
            width, height = img.size
        except Exception:
            pass
        return {
            'width':   width,
            'height':  height,
            'size_kb': round(size_bytes / 1024, 1) if size_bytes else None,
            'ok':      True,
        }
    except Exception:
        return None


def cargar_metadata_excel(meta_path):
    """
    Carga (o crea) el Excel de metadata.
    Retorna (workbook, worksheet, {clave -> fila_idx}, fila_maxima).
    """
    if not os.path.exists(meta_path):
        # Crear nuevo Excel con encabezados
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Metadata'
        for i, h in enumerate(ENCABEZADOS, 1):
            ws.cell(row=1, column=i, value=h)
        wb.save(meta_path)
        return wb, ws, {}, 1

    wb = openpyxl.load_workbook(meta_path)
    ws = wb.active
    indice    = {}
    fila_max  = 1
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        clave_cell = row[0].value
        if clave_cell:
            indice[str(clave_cell).strip()] = idx
            fila_max = idx
    return wb, ws, indice, fila_max


# ── Comando ────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = (
        'Extrae metadatos de videos/imágenes en Drive y los guarda en '
        'MATEO METADATA.xlsx (clave ID__Estilo__Miembro).'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--limit', type=int, default=0,
            help='Maximo de registros a procesar (0=todos)'
        )
        parser.add_argument(
            '--id', type=str, default='',
            help='Procesar solo el registro con este ID'
        )
        parser.add_argument(
            '--force', action='store_true',
            help='Re-extraer metadatos aunque ya existan en el Excel'
        )
        parser.add_argument(
            '--reset', action='store_true',
            help='Borrar MATEO METADATA.xlsx y empezar de cero'
        )

    def handle(self, *args, **options):
        registro_path = os.path.join(settings.BASE_DIR, REGISTRO_EXCEL)
        meta_path     = os.path.join(settings.BASE_DIR, METADATA_EXCEL)
        limite    = options.get('limit', 0)
        id_filtro = options.get('id', '').strip()
        forzar    = options.get('force', False)
        reset     = options.get('reset', False)

        if not os.path.exists(registro_path):
            self.stderr.write(f'[ERROR] Excel no encontrado: {registro_path}')
            return

        # Reset: eliminar el Excel de metadata para empezar de cero
        if reset and os.path.exists(meta_path):
            os.remove(meta_path)
            self.stdout.write('[RESET] Excel de metadata eliminado. Se creara de cero.')

        # Cargar (o crear) el Excel de metadata
        wb_meta, ws_meta, indice_meta, fila_maxima = cargar_metadata_excel(meta_path)

        # Leer el Excel de registro con data_only=False para preservar hipervinculos
        wb_reg = openpyxl.load_workbook(registro_path, data_only=False)
        ws_reg = wb_reg.active

        # Cargar datos de contenido del censo de Fase 1 (mapa, especie, genero, etnia)
        # El censo indexa por ID entero ("1"), el registro usa floats ("1.0") -> normalizar
        def norm_id(id_v):
            """Normaliza IDs numericos: '1.0' -> '1'."""
            try:
                f2 = float(str(id_v).strip())
                if f2 == int(f2):
                    return str(int(f2))
            except (ValueError, TypeError):
                pass
            return str(id_v).strip()

        censo_path = os.path.join(settings.BASE_DIR, CENSO_CSV)
        censo_lookup = {}
        if os.path.exists(censo_path):
            with open(censo_path, mode='r', encoding='utf-8-sig') as fc:
                reader = csv.DictReader(fc)
                for crow in reader:
                    vid_id = norm_id(
                        crow.get('ID DE VIDEO') or crow.get('ID DE VIDEO EQUIPO') or ''
                    )
                    if vid_id:
                        censo_lookup[vid_id] = {
                            'mapa':    str(crow.get('MAPA', '')).strip(),
                            'especie': str(crow.get('ESPECIE', '')).strip(),
                            'genero':  str(crow.get('GENERO', '')).strip(),
                            'etnia':   str(crow.get('ETNIA', '')).strip(),
                        }
            self.stdout.write(f'[INFO] {len(censo_lookup)} registros cargados del censo.csv')
        else:
            self.stdout.write('[WARN] censo.csv no encontrado — mapa/especie/genero/etnia quedaran vacios')

        # Construir lista de tareas (registros pendientes de extracción)
        tareas = []
        for row in ws_reg.iter_rows(min_row=2):
            if not any(c.value for c in row[:4]):
                continue
            miembro  = extraer_celda(row[0])
            id_reg   = extraer_celda(row[1])
            estilo   = extraer_celda(row[3])
            if id_filtro and id_reg != id_filtro:
                continue
            clave = f'{id_reg}__{estilo}__{miembro}'
            # Saltar si ya tiene datos y no se fuerza
            if clave in indice_meta and not forzar:
                continue
            tareas.append({
                'clave':           clave,
                'id_reg':          id_reg,
                'estilo':          estilo,
                'miembro':         miembro,
                'url_imagen':      extraer_celda(row[5], col_idx=5)  if len(row) > 5  else '',
                'url_vid_estil':   extraer_celda(row[7], col_idx=7)  if len(row) > 7  else '',
                'url_vid_orig':    extraer_celda(row[8], col_idx=8)  if len(row) > 8  else '',
            })

        self.stdout.write(f'[INFO] {len(indice_meta)} registros ya en el Excel de metadata')
        self.stdout.write(f'[INFO] {len(tareas)} registros pendientes de extraccion')

        if limite > 0:
            tareas = tareas[:limite]
            self.stdout.write(f'[INFO] Limitando a {len(tareas)} registros (--limit {limite})')

        if not tareas:
            self.stdout.write('[OK] No hay registros nuevos que procesar')
            return

        ok_count = 0
        err_count = 0

        for i, t in enumerate(tareas, start=1):
            self.stdout.write(f'  [{i}/{len(tareas)}] {t["clave"]}')
            ahora = datetime.now().isoformat()

            # ── Video original ─────────────────────────
            fid_orig  = extraer_file_id(t['url_vid_orig'])
            dur_orig  = None
            if fid_orig:
                meta_orig = extraer_metadata_video(fid_orig)
                if meta_orig and meta_orig.get('ok'):
                    dur_orig = meta_orig.get('duration_s')

            # ── Video estilizado ───────────────────────
            fid_estil = extraer_file_id(t['url_vid_estil'])
            dur_estil = fps_estil = res_estil = size_estil = None
            if fid_estil:
                meta_estil = extraer_metadata_video(fid_estil)
                if meta_estil and meta_estil.get('ok'):
                    dur_estil  = meta_estil.get('duration_s')
                    fps_estil  = meta_estil.get('fps')
                    res_estil  = meta_estil.get('resolution')
                    size_estil = meta_estil.get('size_mb')

            # ── Imagen generada ────────────────────────
            fid_img = extraer_file_id(t['url_imagen'])
            img_w = img_h = img_kb = None
            if fid_img:
                meta_img = extraer_metadata_imagen(fid_img)
                if meta_img and meta_img.get('ok'):
                    img_w  = meta_img.get('width')
                    img_h  = meta_img.get('height')
                    img_kb = meta_img.get('size_kb')

            # Campos de contenido del video original (del censo de Fase 1)
            censo_data = censo_lookup.get(norm_id(t['id_reg']), {})
            t_mapa    = censo_data.get('mapa', '')
            t_especie = censo_data.get('especie', '')
            t_genero  = censo_data.get('genero', '')
            t_etnia   = censo_data.get('etnia', '')

            # Determinar si se obtuvo al menos un dato de metadata
            meta_ok = any(v is not None for v in [dur_orig, dur_estil, img_w])

            fila_vals = [
                t['clave'], t['id_reg'], t['estilo'], t['miembro'],
                dur_orig, dur_estil, fps_estil, res_estil, size_estil,
                img_w, img_h, img_kb,
                meta_ok, ahora,
                t_mapa, t_especie, t_genero, t_etnia,
            ]

            # Insertar nueva fila o actualizar la existente
            if t['clave'] in indice_meta:
                fila_idx = indice_meta[t['clave']]
            else:
                fila_maxima += 1
                fila_idx = fila_maxima
                indice_meta[t['clave']] = fila_idx

            for col_i, val in enumerate(fila_vals, start=1):
                ws_meta.cell(row=fila_idx, column=col_i, value=val)

            if meta_ok:
                self.stdout.write(
                    f'    [OK] dur_orig={dur_orig}s  dur_estil={dur_estil}s  '
                    f'fps={fps_estil}  res={res_estil}'
                )
                ok_count += 1
            else:
                self.stdout.write('    [WARN] No se pudo extraer metadata')
                err_count += 1

            # Guardar parcialmente cada 10 registros (por si se interrumpe)
            if i % 10 == 0:
                wb_meta.save(meta_path)
                self.stdout.write('  [SAVE] Excel de metadata guardado parcialmente')

        # Guardar el Excel final
        wb_meta.save(meta_path)
        self.stdout.write(
            f'\n[OK] Extraccion completada:'
            f'\n   Exitosos:  {ok_count}'
            f'\n   Errores:   {err_count}'
            f'\n   Excel guardado en: {meta_path}'
        )
