"""
Comando Django: python manage.py sync_registro
Sincroniza el Excel local con los datos del Google Sheets publico.

Descarga en formato XLSX para preservar hipervinculos de Google Drive.
Usa clave compuesta (ID + Estilo) porque el mismo video puede tener
multiples versiones estilizadas (Anime, Cartoon, Lego, Ciberpunk).
Las columnas de link (F, H, I, M, O) extraen la URL del hipervinculo;
las columnas de texto (prompt, ACEPTADO, etc.) usan siempre el valor de texto.
"""
from django.core.management.base import BaseCommand
import requests
import urllib3
import os
import io
import re
import openpyxl
from django.conf import settings

# Suprimir advertencia de SSL no verificado (red corporativa con cert self-signed)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Exportar como XLSX para preservar hipervinculos (CSV los pierde)
SHEETS_XLSX_URL = (
    'https://docs.google.com/spreadsheets/d/'
    '1Ga5zMekIlVjHKhxkfGBIhAiCh0H3zGYf0TOoMkoctj4'
    '/export?format=xlsx&gid=2173056'
)

NUM_COLS = 15  # columnas A-O que nos interesan

# Indices (0-based) de columnas que contienen hipervinculos de Google Drive:
#   F=5 (imagen), H=7 (video), I=8 (video original), M=12 (imagen arreglo), O=14 (video arreglo)
COLS_LINK = {5, 7, 8, 12, 14}


def extraer_celda(cell, col_idx=None):
    """
    Extrae el valor util de una celda respetando el tipo de columna:

    - Columnas de LINK (F, H, I, M, O): prioriza hipervinculo > formula HYPERLINK > URL directa.
      Estas columnas contienen los links de Google Drive de los archivos.
    - Columnas de TEXTO (prompt, ACEPTADO, Miembro, etc.): usa SIEMPRE el valor de texto.
      Google Sheets puede adjuntar hipervinculos accidentales a celdas de texto;
      si extrajéramos el hipervinculo obtendriamos una URL en lugar del prompt.
    """
    val = cell.value
    texto = str(val).strip() if val is not None else ''
    if texto.lower() == 'none':
        texto = ''

    if col_idx in COLS_LINK:
        # Columna de link: intentar obtener la URL real del hipervinculo
        if cell.hyperlink and cell.hyperlink.target:
            return str(cell.hyperlink.target).strip()
        # Parsear formula HYPERLINK literal: =HYPERLINK("url","texto") o =HYPERLINK("url")
        m = re.match(r'=HYPERLINK\("([^"]+)"', texto, re.IGNORECASE)
        if m:
            return m.group(1)
        # Si el valor ya es una URL directa
        if texto.lower().startswith('http'):
            return texto

    # Para columnas de texto: retornar siempre el valor de texto de la celda
    return texto


def hacer_clave(vals):
    """
    Clave unica compuesta: ID + Estilo + Miembro.
    Necesaria porque:
    - El mismo ID puede tener multiples estilos (Anime, Cartoon, Lego, Ciberpunk).
    - El mismo ID + Estilo puede tener multiples miembros (Wilson y Rodrigo
      pueden tener ambos una version Lego del video ID=1).
    """
    id_v    = str(vals[1]).strip()
    estilo  = str(vals[3]).strip()
    miembro = str(vals[0]).strip()
    return f"{id_v}__{estilo}__{miembro}"


class Command(BaseCommand):
    help = 'Sincroniza el Excel local con el Google Sheets (XLSX, clave ID+Estilo, preserva hipervinculos)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Solo muestra los cambios sin modificar el Excel'
        )
        parser.add_argument(
            '--force', action='store_true',
            help='Sobrescribe todas las filas aunque no haya diferencias detectadas'
        )

    def handle(self, *args, **options):
        excel_path = os.path.join(settings.BASE_DIR, 'MATEO REGISTRO PARAMETROS.xlsx')
        dry_run = options.get('dry_run', False)
        force   = options.get('force', False)

        self.stdout.write('[SYNC] Descargando Google Sheets como XLSX...')

        # -- 1. Descargar el XLSX del Sheets --
        try:
            # verify=False necesario por certificado self-signed en red corporativa
            resp = requests.get(SHEETS_XLSX_URL, timeout=60, verify=False)
            resp.raise_for_status()
        except Exception as e:
            self.stderr.write(f'[ERROR] No se pudo descargar el Sheets: {e}')
            return

        # data_only=False para leer valores reales e hipervinculos
        try:
            wb_sheets = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
            ws_sheets = wb_sheets.active
        except Exception as e:
            self.stderr.write(f'[ERROR] No se pudo leer el XLSX descargado: {e}')
            return

        # Construir diccionario {clave_ID_Estilo -> [col1..col15]}
        # Para cada celda se pasa el col_idx para distinguir texto vs link
        sheets_por_clave = {}
        for row in ws_sheets.iter_rows(min_row=2):
            vals = [
                extraer_celda(row[i], col_idx=i) if i < len(row) else ''
                for i in range(NUM_COLS)
            ]
            if not any(vals[:4]):
                continue
            if not vals[1].strip():
                continue
            clave = hacer_clave(vals)
            sheets_por_clave[clave] = vals

        self.stdout.write(f'[OK] {len(sheets_por_clave)} registros unicos (ID+Estilo) en el Sheets')

        # -- 2. Cargar el Excel local --
        if not os.path.exists(excel_path):
            self.stderr.write(f'[ERROR] Excel no encontrado: {excel_path}')
            return

        wb_local = openpyxl.load_workbook(excel_path)
        ws_local = wb_local.active

        # Construir indice {clave_ID_Estilo -> numero_de_fila}
        excel_por_clave = {}
        fila_maxima = 1
        for idx, row in enumerate(ws_local.iter_rows(min_row=2), start=2):
            if len(row) < 4:
                continue
            vals = [extraer_celda(row[i], col_idx=i) if i < len(row) else ''
                    for i in range(NUM_COLS)]
            if not any(vals[:4]) or not vals[1].strip():
                continue
            clave = hacer_clave(vals)
            excel_por_clave[clave] = idx
            fila_maxima = idx

        self.stdout.write(f'[INFO] {len(excel_por_clave)} registros unicos (ID+Estilo) en el Excel local')

        # -- 3. Detectar diferencias --
        nuevas       = []   # filas que estan en Sheets pero no en local
        actualizadas = []   # (fila_idx, fila_sheets) donde los datos difieren

        for clave, fila_s in sheets_por_clave.items():
            if clave not in excel_por_clave:
                nuevas.append(fila_s)
            else:
                fila_idx   = excel_por_clave[clave]
                fila_excel = [
                    extraer_celda(ws_local.cell(fila_idx, c + 1), col_idx=c)
                    for c in range(NUM_COLS)
                ]
                if force or fila_excel != fila_s:
                    actualizadas.append((fila_idx, fila_s))

        self.stdout.write(
            f'\n[RESUMEN] Cambios detectados:'
            f'\n   Filas nuevas (ID+Estilo no existe en local): {len(nuevas)}'
            f'\n   Filas actualizadas (datos difieren):          {len(actualizadas)}'
        )

        if dry_run:
            # Muestra de los primeros cambios para verificar
            if nuevas:
                self.stdout.write('\n[DRY-RUN] Primeras 3 filas NUEVAS:')
                for fila in nuevas[:3]:
                    self.stdout.write(
                        f'  Clave={hacer_clave(fila)!r}: '
                        f'Miembro={fila[0]!r}, Prod={fila[2]!r}, Estilo={fila[3]!r}'
                    )
            if actualizadas:
                self.stdout.write('\n[DRY-RUN] Primeras 3 filas con CAMBIOS:')
                for fila_idx, fila_s in actualizadas[:3]:
                    self.stdout.write(
                        f'  Clave={hacer_clave(fila_s)!r} fila={fila_idx}: '
                        f'Miembro={fila_s[0]!r}, prompt={fila_s[4][:60]!r}'
                    )
            self.stdout.write('\n[DRY-RUN] No se realizaron cambios')
            return

        # -- 4. Aplicar cambios al Excel local --
        cambios_totales = 0

        # Agregar filas nuevas al final del Excel
        for fila in nuevas:
            fila_maxima += 1
            for col_idx, valor in enumerate(fila, start=1):
                ws_local.cell(row=fila_maxima, column=col_idx, value=valor)
            cambios_totales += 1

        # Actualizar filas existentes con datos del Sheets
        for fila_idx, fila_s in actualizadas:
            for col_idx, valor in enumerate(fila_s, start=1):
                ws_local.cell(row=fila_idx, column=col_idx, value=valor)
            cambios_totales += 1

        if cambios_totales > 0:
            wb_local.save(excel_path)
            self.stdout.write(f'\n[OK] Excel actualizado: {cambios_totales} filas modificadas')
        else:
            self.stdout.write('\n[OK] El Excel ya estaba al dia, no se realizaron cambios')
